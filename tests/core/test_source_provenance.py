"""Tests for the source-provenance columns (source_chunk, source_sentence, conf_score)."""

import io

from openpyxl import load_workbook

from core.jama_hierarchy import COLS, build_hierarchy_workbook_bytes
from core.source_matching import best_source_sentence, sanitize_excel_text, split_into_sentences


def test_sanitize_excel_text_strips_leading_equals():
    assert sanitize_excel_text("= The system shall operate") == "The system shall operate"
    assert sanitize_excel_text("===value") == "value"
    assert sanitize_excel_text("The = operator is used") == "The = operator is used"
    assert sanitize_excel_text("") == ""
    assert sanitize_excel_text("   = padded") == "padded"


def test_best_source_sentence_strips_leading_equals_from_match():
    chunk = "= The system shall seat eight passengers. Speed capped at 120 km/h."

    def scorer(description, sentences):
        return [1.0 if sentences[0].startswith("=") else 0.0 for _ in sentences]

    sentence, score = best_source_sentence(
        "The system shall seat eight passengers.", chunk, scorer=scorer
    )
    assert sentence == "The system shall seat eight passengers."
    assert not sentence.startswith("=")


def test_hierarchy_workbook_source_sentence_never_starts_with_equals():
    req = {
        "Name": "Formula-like Source",
        "Description": "The system shall equal five.",
        "VerificationMethod": "Test",
        "RequirementType": "Functional",
        "Tags": [],
        "DocumentRequirementID": "#99",
        "source_chunk": "= Section intro\n= The system shall equal five.",
        "source_sentence": "= The system shall equal five.",
        "conf_score": 0.88,
    }
    chunk_results = [(["4 Requirements"], [req])]

    xlsx_bytes = build_hierarchy_workbook_bytes(chunk_results)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Jama_Hierarchy"]

    sentence_col = COLS.index("source_sentence") + 1
    chunk_col = COLS.index("source_chunk") + 1
    id_col = COLS.index("DocumentRequirementID") + 1
    req_row = next(
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=id_col).value == "#99"
    )

    source_sentence = ws.cell(row=req_row, column=sentence_col).value
    source_chunk = ws.cell(row=req_row, column=chunk_col).value

    assert source_sentence == "The system shall equal five."
    assert not str(source_sentence).startswith("=")
    assert not str(source_chunk).startswith("=")
    assert source_chunk == "Section intro\n= The system shall equal five."


def test_split_into_sentences_line_and_punctuation_aware():
    text = "The system shall do A. It shall also do B!\n- A bullet requirement\n"
    sentences = split_into_sentences(text)
    assert sentences == [
        "The system shall do A.",
        "It shall also do B!",
        "- A bullet requirement",
    ]


def test_split_into_sentences_empty():
    assert split_into_sentences("") == []
    assert split_into_sentences("   \n  ") == []


def test_best_source_sentence_uses_injected_scorer():
    chunk = "The vehicle shall reach 120 km/h. The cabin shall seat eight passengers."
    # Injected scorer: reward the sentence mentioning passengers.
    def scorer(description, sentences):
        return [1.0 if "passengers" in s else 0.1 for s in sentences]

    sentence, score = best_source_sentence(
        "The system shall seat eight occupants.", chunk, scorer=scorer
    )
    assert sentence == "The cabin shall seat eight passengers."
    assert score == 1.0


def test_best_source_sentence_no_sentences():
    assert best_source_sentence("anything", "") == ("", 0.0)


def test_best_source_sentence_lexical_fallback_is_deterministic():
    chunk = "Speed is limited to 120 km/h. The system shall seat eight passengers."
    # No scorer -> embedding attempt then lexical fallback; must still return a
    # sentence from the chunk and a numeric score.
    sentence, score = best_source_sentence(
        "The system shall seat eight passengers.", chunk
    )
    assert sentence in split_into_sentences(chunk)
    assert isinstance(score, float)


def test_hierarchy_workbook_includes_provenance_columns_and_values():
    req = {
        "Name": "Passenger Capacity",
        "Description": "The system shall seat eight passengers.",
        "VerificationMethod": "Test",
        "RequirementType": "Functional",
        "Tags": ["TBD"],
        "DocumentRequirementID": "#12",
        "source_chunk": "The system shall seat eight passengers. Speed capped at 120 km/h.",
        "source_sentence": "The system shall seat eight passengers.",
        "conf_score": 0.9123,
    }
    chunk_results = [(["4 Requirements", "4.1 Capacity"], [req])]

    xlsx_bytes = build_hierarchy_workbook_bytes(chunk_results)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Jama_Hierarchy"]

    header = [c.value for c in ws[1]]
    assert header == COLS
    assert header[-3:] == ["source_chunk", "source_sentence", "conf_score"]

    # Find the requirement row (the one carrying the DocumentRequirementID).
    id_col = COLS.index("DocumentRequirementID") + 1
    req_row = next(
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=id_col).value == "#12"
    )
    chunk_col = COLS.index("source_chunk") + 1
    sentence_col = COLS.index("source_sentence") + 1
    conf_col = COLS.index("conf_score") + 1

    assert ws.cell(row=req_row, column=chunk_col).value == req["source_chunk"]
    assert ws.cell(row=req_row, column=sentence_col).value == req["source_sentence"]
    assert ws.cell(row=req_row, column=conf_col).value == req["conf_score"]


def test_hierarchy_section_rows_leave_provenance_blank():
    # A heading with no requirements should still render with empty provenance cells.
    chunk_results = [(["4 Requirements"], [])]
    xlsx_bytes = build_hierarchy_workbook_bytes(chunk_results)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Jama_Hierarchy"]

    chunk_col = COLS.index("source_chunk") + 1
    conf_col = COLS.index("conf_score") + 1
    # Row 2 is the section row.
    assert ws.cell(row=2, column=chunk_col).value in (None, "")
    assert ws.cell(row=2, column=conf_col).value in (None, "")
