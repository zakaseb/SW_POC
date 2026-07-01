"""
jama_hierarchy.py
-----------------
Build a plain, indentation-only "Jama_Hierarchy" sheet driven by the HEADING
paths captured on requirement chunks (Docling `meta.headings`), rather than a
hard-coded section tree.

Input shape (`chunk_results`):
    list of (headings_path, requirements) in document order, where
      headings_path : list[str]   heading chain root->leaf for the chunk
      requirements  : list[dict]  parsed requirement dicts for that chunk
                                  (keys: Name, Description, VerificationMethod,
                                   Tags, RequirementType, DocumentRequirementID)

Layout (mirrors the attached hierarchy_sheet.py style):
  - native Excel indent on the Name column; a Level column shows the depth
  - each unique heading-path prefix emitted once as a section row (in document
    order of first appearance)
  - requirements nested one level deeper than their leaf heading
  - a heading with no requirements is simply a section row with empty req cols
"""

from __future__ import annotations
import regex as re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

COLS = [
    "Level",
    "Name",
    "Description",
    "VerificationMethod",
    "RequirementType",
    "Tags",
    "DocumentRequirementID",
    "source_chunk",
    "source_sentence",
    "conf_score",
]
COL_WIDTHS = {1: 8, 2: 60, 3: 70, 4: 22, 5: 22, 6: 18, 7: 22, 8: 80, 9: 70, 10: 12}

PLAIN_FONT  = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True)
CENTER      = Alignment(horizontal="center", vertical="center")
_SEP = r"[\s\u00a0\u200b\u200c\u200d\u2060\ufeff]"
_SECTION_NUM_RE = re.compile(rf"^\s*\d+(?:\.\d+)*\.?{_SEP}+")

def _tags_to_str(value):
    return ", ".join(value) if isinstance(value, list) else (value or "")


def _plain_row(ws, row_idx, values, height=15, indent=0):
    ws.row_dimensions[row_idx].height = height
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font = PLAIN_FONT
        c.alignment = Alignment(
            horizontal="left",
            wrap_text=True,
            vertical="top",
            indent=(indent if col == 2 else 0),   # indent the Name column only
        )


def _section_row(ws, row_idx, level, name):
    _plain_row(ws, row_idx, [str(level), name, "", "", "", "", "", "", "", ""],
               height=15, indent=level)


def _requirement_row(ws, row_idx, level, req):
    _plain_row(
        ws,
        row_idx,
        [
            str(level),
            req.get("Name", ""),
            req.get("Description", ""),
            req.get("VerificationMethod", ""),
            req.get("RequirementType", ""),
            _tags_to_str(req.get("Tags", "")),
            req.get("DocumentRequirementID", ""),
            req.get("source_chunk", ""),
            req.get("source_sentence", ""),
            req.get("conf_score", ""),
        ],
        height=30,
        indent=level,
    )


def _write_header(ws):
    for j, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[1].height = 18

def _strip_section_number(name: str) -> str:
    """'2.1.2.2 BOOT State' -> 'BOOT State'; leaves un-numbered names untouched."""
    cleaned = _SECTION_NUM_RE.sub("", name).strip()
    return cleaned or name        # fall back to original if stripping leaves nothing

def _build_tree(ws, chunk_results):
    row = 2
    offset = 0

    emitted_prefixes: set[tuple] = set()
    for headings_path, requirements in chunk_results:
        path = list(headings_path) if headings_path else ["(no heading)"]
        if path == ["(no heading)"]:
            continue
        # emit each unique heading-path prefix once, in first-seen (document) order
        for depth in range(len(path)):
            prefix = tuple(path[: depth + 1])
            if prefix not in emitted_prefixes:
                emitted_prefixes.add(prefix)
                _section_row(ws, row, depth + offset, _strip_section_number(path[depth]))
                row += 1

        # requirements as children, one level below the leaf heading
        child_level = len(path) + offset
        for req in requirements:
            _requirement_row(ws, row, child_level, req)
            row += 1

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def build_hierarchy_sheet_from_chunks(
    wb: Workbook,
    chunk_results,
    sheet_name: str = "Jama_Hierarchy",
):
    """Add a plain indentation-only hierarchy sheet to `wb`. Returns the sheet."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    _write_header(ws)
    _build_tree(ws, chunk_results)
    return ws


def build_hierarchy_workbook_bytes(
    chunk_results,
) -> bytes:
    """Convenience: build a one-sheet workbook and return xlsx bytes."""
    import io
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    build_hierarchy_sheet_from_chunks(wb, chunk_results)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()